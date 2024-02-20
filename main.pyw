from tkinter import Tk, Frame, Button, Label, ttk, PhotoImage, StringVar, Entry
from comunicacion_serial import Comunicacion
from openpyxl import Workbook
from datetime import datetime
import threading
import collections

class Grafica(Frame):
   def __init__(self,master, *args):
      super().__init__(master, *args)

      self.datos_placa = Comunicacion()
      self.datos_placa.puertos_disponibles()

      self.nombre_documento = StringVar()
      self.factor_1 = StringVar() # constante para el cuadalimetro 1
      self.factor_2 = StringVar() # constante para el cuadalimetro 2

      self.datos = 0.0 #variable para alacenar lo ingreasado por uart

      self.sensor = 0 # Espacio de memoria para almacenar el sensor activado
      self.valor_adc = 0 # Espacio de memoria para almacenar el valor adc
      
      self.libro = Workbook() # Crear un nuevo libro de trabajo (workbook)
      self.sheet = self.libro.active # Seleccionar la hoja activa (por defecto la primera hoja)

      self.fila_sensor_1 = 0 # Define los espacios para almacenar la fila
      self.fila_sensor_2 = 0
      self.fila_sensor_3 = 0
      self.fila_sensor_4 = 0

      self.aux_1 = 0 # Variable axiliar para guardar fila momentaneamente
      self.aux_2 = 0 # Variable axiliar para guardar columnas  momentaneos
      #self.columna = 0 #define el espaicio para almacenar la columna

      self.tiempo_anterior_1 = datetime.now() # variable para almacenar el tiempo anterior
      self.tiempo_anterior_2 = datetime.now()
      self.tiempo_anterior_3 = datetime.now()
      self.tiempo_anterior_4 = datetime.now()

      self.tiempo_1 = datetime.now() #variable para guardar el tiempo actual
      self.tiempo_2 = datetime.now() #variable para guardar el tiempo actual
      self.tiempo_3 = datetime.now() #variable para guardar el tiempo actual
      self.tiempo_4 = datetime.now() #variable para guardar el tiempo actual

      self.delta_1 = 0 # variable para guardar la variacion temporal entre muestras
      self.delta_2 = 0 # variable para guardar la variacion temporal entre muestras
      self.delta_3 = 0 # variable para guardar la variacion temporal entre muestras
      self.delta_4 = 0 # variable para guardar la variacion temporal entre muestras

      self.valor_adc_1 = 0 # Almacena el ultimo valor adc
      self.valor_adc_2 = 0
      self.valor_adc_3 = 0
      self.valor_adc_4 = 0

      self.fecha = datetime.now() # variable para guardar la fecha
      self.hora_aux = "" #variable para almacenar la hora formateada
      self.delta_aux = "" #variable para almacenar la variacion de teimpo

      self.tiempo_actual_fomateado = ["none", "none", "none", "none"] # para guardar los tiempos formateados
      self.tiempo_anterior_fomateado = ["none", "none", "none", "none"]
      self.caudal = [0, 0 , 0, 0] # guarda los caudales
      self.factor = [0, 0] # guarda los factores

      self.datos_placa.recibida.clear() #Se limpia el evento del hilo principal por primera vez
      self.widgets()


   def fn_logica(self):

      temporal = 0
      caudal_aux = 0

      if self.sensor == "1":
         self.fila_sensor_1 += 1
         self.aux_1 = self.fila_sensor_1
         self.aux_2 = 1 # donde inicia la columna de cada sensor
         self.tiempo_anterior_1 = self.tiempo_1 # Gurada el tiempo anterior
         self.tiempo_1 = datetime.now() # obtien la hora actual
         temporal = self.tiempo_1 - self.tiempo_anterior_1 # calcula la variacion temporal
         self.delta_1 = round(float(temporal.total_seconds()), 4)
         self.delta_aux = self.delta_1 # almacena momentaneamente el valor de delta para guardarlo
         self.hora_aux = self.tiempo_1.strftime("%H:%M:%S") # formatea el valor para almacenar hoja de calculo
         self.valor_adc_1 = self.valor_adc # para mostra el valor por pantalla
         self.tiempo_anterior_fomateado[0] = self.tiempo_actual_fomateado[0]
         self.tiempo_actual_fomateado[0] = self.hora_aux # guarda la hora formateada
         self.caudal[0] = round(self.factor[0] / self.delta_1, 4) #calcula el caudal con la variacion temporal y el factor
         caudal_aux = self.caudal[0] # auxiliar para almacenar en el documento
      elif self.sensor == "2":
         self.fila_sensor_2 += 1
         self.aux_1 = self.fila_sensor_2
         self.aux_2 = 1 + self.n
         self.tiempo_anterior_2 = self.tiempo_2 # Gurada el tiempo anterior
         self.tiempo_2 = datetime.now() # obtien la hora actual
         temporal = self.tiempo_2 - self.tiempo_anterior_2 # calcula la variacion temporal
         self.delta_2 = round(float(temporal.total_seconds()), 4)
         self.delta_aux = self.delta_2 # almacena momentaneamente el valor de delta para guardarlo
         self.hora_aux = self.tiempo_2.strftime("%H:%M:%S") # formatea el valor para almacenar hoja de calculo
         self.valor_adc_2 = self.valor_adc # para mostra el valor por pantalla
         self.tiempo_anterior_fomateado[1] = self.tiempo_actual_fomateado[1]
         self.tiempo_actual_fomateado[1] = self.hora_aux
         self.caudal[1] = round(self.factor[0] / self.delta_2, 4) #calcula el caudal con la variacion temporal y el factor
         caudal_aux = self.caudal[1]
      elif self.sensor == "3":
         self.fila_sensor_3 += 1
         self.aux_1 = self.fila_sensor_3
         self.aux_2 = 1 + 2*self.n
         self.tiempo_anterior_3 = self.tiempo_3 # Gurada el tiempo anterior
         self.tiempo_3 = datetime.now() # obtien la hora actual
         temporal = self.tiempo_3 - self.tiempo_anterior_3 # calcula la variacion temporal
         self.delta_3 = round(float(temporal.total_seconds()),4)
         self.delta_aux = self.delta_3 # almacena momentaneamente el valor de delta para guardarlo
         self.hora_aux = self.tiempo_3.strftime("%H:%M:%S") # formatea el valor para almacenar hoja de calculo
         self.valor_adc_3 = self.valor_adc # para mostra el valor por pantalla
         self.tiempo_anterior_fomateado[2] = self.tiempo_actual_fomateado[2]
         self.tiempo_actual_fomateado[2] = self.hora_aux
         self.caudal[2] = round(self.factor[1] / self.delta_3, 4) #calcula el caudal con la variacion temporal y el factor
         caudal_aux = self.caudal[2]
      elif self.sensor == "4":
         self.fila_sensor_4 += 1
         self.aux_1 = self.fila_sensor_4
         self.aux_2 = 1 + 3 * self.n
         self.tiempo_anterior_4 = self.tiempo_4 # Gurada el tiempo anterior
         self.tiempo_4 = datetime.now() # obtien la hora actual
         temporal = self.tiempo_4 - self.tiempo_anterior_4 # calcula la variacion temporal
         self.delta_4 = round(float(temporal.total_seconds()), 4)
         self.delta_aux = self.delta_4 # almacena momentaneamente el valor de delta para guardarlo
         self.hora_aux = self.tiempo_4.strftime("%H:%M:%S") # formatea el valor para almacenar hoja de calculo
         self.valor_adc_4 = self.valor_adc # para mostra el valor por pantalla
         self.tiempo_anterior_fomateado[3] = self.tiempo_actual_fomateado[3]
         self.tiempo_actual_fomateado[3] = self.hora_aux
         self.caudal[3] = round(self.factor[1] / self.delta_4, 4) #calcula el caudal con la variacion temporal y el factor
         caudal_aux = self.caudal[3]

      # guarda la hora
      self.sheet.cell(row=self.aux_1, column=self.aux_2, value=self.hora_aux)
      # guarda la variacion temporal
      self.aux_2 += 1
      self.sheet.cell(row=self.aux_1, column=self.aux_2, value=self.delta_aux)
      # guarda el el valor adc
      self.aux_2 += 1
      self.sheet.cell(row=self.aux_1, column=self.aux_2, value=self.valor_adc)
      # guarda el caudal
      self.aux_2 += 1
      self.sheet.cell(row=self.aux_1, column=self.aux_2, value=caudal_aux)

   ## Actualiza el valor de las etiquetas de la pantalla  
   def ColocarValores(self):

      # sensor 1
      self.valor_actual_1_t.config(text=self.valor_adc_1)
      self.tiempo_1_t.config(text=self.tiempo_actual_fomateado[0])
      self.tiempo_anterior_1_t.config(text=self.tiempo_anterior_fomateado[0])
      self.delta_1_t.config(text=self.delta_1)
      self.caudal_1_t.config(text=self.caudal[0])

      # sensor 2
      self.valor_actual_2_t.config(text=self.valor_adc_2)
      self.tiempo_2_t.config(text=self.tiempo_actual_fomateado[1])
      self.tiempo_anterior_2_t.config(text=self.tiempo_anterior_fomateado[1])
      self.delta_2_t.config(text=self.delta_2)
      self.caudal_2_t.config(text=self.caudal[1])

      # sensor 3
      self.valor_actual_3_t.config(text=self.valor_adc_3)
      self.tiempo_3_t.config(text=self.tiempo_actual_fomateado[2])
      self.tiempo_anterior_3_t.config(text=self.tiempo_anterior_fomateado[2])
      self.delta_3_t.config(text=self.delta_3)
      self.caudal_3_t.config(text=self.caudal[2])

      # sensor 2
      self.valor_actual_4_t.config(text=self.valor_adc_4)
      self.tiempo_4_t.config(text=self.tiempo_actual_fomateado[3])
      self.tiempo_anterior_4_t.config(text=self.tiempo_anterior_fomateado[3])
      self.delta_4_t.config(text=self.delta_4)
      self.caudal_4_t.config(text=self.caudal[3])

   def HiloPrincipal(self):

      ## Crea los titulos en el libro
      self.sheet['A1'] = 'Fecha'
      self.sheet['B1'] =self.fecha.strftime("%Y-%m-%d")

      self.sheet['A2'] = 'Sensor 1'
      self.sheet['A3'] = 'Hora'
      self.sheet['B3'] = 'Variacion de tiempo'
      self.sheet['C3'] = 'Ciclo de trabajo %'
      self.sheet['D3'] = 'Caudal'

      self.sheet['E2'] = 'Sensor 2'
      self.sheet['E3'] = 'Hora'
      self.sheet['F3'] = 'Variacion de tiempo'
      self.sheet['G3'] = 'Ciclo de trabajo %'
      self.sheet['H3'] = 'Caudal'

      self.sheet['I2'] = 'Sensor 3'
      self.sheet['I3'] = 'Hora'
      self.sheet['J3'] = 'Variacion de tiempo'
      self.sheet['K3'] = 'Ciclo de trabajo %'
      self.sheet['L3'] = 'Caudal'

      self.sheet['M2'] = 'Sensor 4'
      self.sheet['M3'] = 'Hora'
      self.sheet['N3'] = 'Variacion de tiempo'
      self.sheet['O3'] = 'Ciclo de trabajo %'
      self.sheet['P3'] = 'Caudal'

      self.fila_sensor_1 = 3
      self.fila_sensor_2 = 3
      self.fila_sensor_3 = 3
      self.fila_sensor_4 = 3
      self.n = 4 # define el numero de casillas de variables a guardar en el documento
 
      while True:
         self.datos_placa.recibida.clear() #Limpia el evento
         self.datos_placa.recibida.wait() #Espera a que se active el evento nuevamente

         self.datos = self.datos_placa.datos_recibidos
         print(self.datos)
         dato = self.datos.split(" ")

         try:
            self.sensor = dato[0]
            self.valor_adc = round(((float(dato[1])-204) / 819) * 100, 4) # se obtiene el ciclo de trabajo 
         except:
            pass

         if self.sensor:
            self.fn_logica()
            self.ColocarValores()
         

   def CrearHilo(self):
      
      hilo_principal = threading.Thread(target=self.HiloPrincipal)
      hilo_principal.setDaemon(1)
      hilo_principal.start() #Inicia el hilo principal

   def widgets(self):
      ## define los frames en lo que se divide la interface grafica
      frame = Frame(self.master, bg='#090818', bd=4)
      frame.grid(column=0, row=0, sticky='nsew')
      frame0 = Frame(self.master, bg='#040208', bd=4)             
      frame0.grid(column=1, row=0, sticky='nsew')
      frame1 = Frame(self.master, bg='#091548', bd=4)
      frame1.grid(column=0, row=1, sticky='nsew')
      frame2 = Frame(self.master, bg='#091708', bd=4)              
      frame2.grid(column=1, row=1, sticky='nsew')
      frame3 = Frame(self.master, bg='#093908', bd=4)             
      frame3.grid(column=0, row=2, sticky='nsew')
      frame4 = Frame(self.master, bg='#091408', bd=4)
      frame4.grid(column=1, row=2, sticky='nsew')
      frame5 = Frame(self.master, bg='#090405', bd=4)
      frame5.grid(column=0, row=3, sticky='nsew')
      frame6 = Frame(self.master, bg='#050401', bd=4)
      frame6.grid(column=1, row=3, sticky='nsew')


      ## establece los tamaños relatitivos de ecpancion de las columas
      self.master.columnconfigure(0, weight=1)
      self.master.columnconfigure(1, weight=3)
      ## establece los tamaños relatitivos de ecpancion de las filas
      self.master.rowconfigure(0, weight=3)
      self.master.rowconfigure(1, weight=2)
      self.master.rowconfigure(2, weight=1)
      self.master.rowconfigure(3, weight=1)

      ## define etiquetas de referencia
      Label(frame, text='Tomar valores', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)
      Label(frame1, text='Guardar Documento', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)
      Label(frame3, text='CONECTAR', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)
      Label(frame4, text='Puertos COM', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)
      Label(frame5, text='DESCONECTAR', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)
      Label(frame6, text='Baudrates', bg='#090808', fg='white', font=('Arial', 12, 'bold')).pack(padx=5, expand=1)

      ##############################################################################################################################

      ## define el cuadro para nombrar el documento del frame 2
      Label(frame2, text='Nombre del Documento:', bg='#090808', fg='white', font=('Arial', 11, 'bold')).grid(row=3, column=0, padx=5, pady=5, sticky='nsew')
      nombre_doc=Entry(frame2, textvariable=self.nombre_documento, font=('Arial', 12, 'bold'))
      nombre_doc.insert(0, "Documento_1") # valor por defecto 
      nombre_doc.grid(row=3, column=1, padx=5, pady=5, sticky='nsew')

      ## define el cuadro para ingresar las constante 1 del frame 2
      Label(frame2, text='Constante 1', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=0, padx=5, pady=5, sticky='nsew')
      self.constante_1_e=Entry(frame2, textvariable=self.factor_1, font=('Arial', 12, 'bold'))
      self.constante_1_e.insert(0, "1") # valor por defecto 
      self.constante_1_e.grid(row=1, column=0, padx=5, pady=5, sticky='nsew')

      ## define el cuadro para ingresar las constante 2 del frame 2
      Label(frame2, text='Constante 2', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=2, padx=5, pady=5, sticky='nsew')
      self.constante_2_e=Entry(frame2, textvariable=self.factor_2, font=('Arial', 12, 'bold'))
      self.constante_2_e.insert(0, "2") # valor por defecto 
      self.constante_2_e.grid(row=1, column=2, padx=5, pady=5, sticky='nsew')

      ## configuracion relativa del los tamaños dentro del frame 0
      frame2.columnconfigure(0,weight=1)
      frame2.columnconfigure(1,weight=1)
      frame2.columnconfigure(2,weight=1)
      frame2.rowconfigure(0, weight=1)
      frame2.rowconfigure(1, weight=1)
      frame2.rowconfigure(2, weight=1)
      frame2.rowconfigure(3, weight=1)

      ##############################################################################################################################

      ## define una variable para mostrar el dato por pantalla del Frame 0
      Label(frame0, text='Ciclo de trabajo %', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=1, column=0, padx=5, pady=5, sticky='nsew')
      Label(frame0, text='Hora Actual', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=2, column=0, padx=5, pady=5, sticky='nsew')
      Label(frame0, text='Hora Anterior', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=3, column=0, padx=5, pady=5, sticky='nsew')
      Label(frame0, text='Tiempo diferencial [Seg]', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=4, column=0, padx=5, pady=5, sticky='nsew')
      Label(frame0, text='Caudal [ ]', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=5, column=0, padx=5, pady=5, sticky='nsew')


      Label(frame0, text='Sensor 1', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=1, padx=5, pady=5, sticky='nsew')
      self.valor_actual_1_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.valor_actual_1_t.grid(row=1, column=1, padx=5, pady=5)
      self.tiempo_1_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_1_t.grid(row=2, column=1, padx=5, pady=5)
      self.tiempo_anterior_1_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_anterior_1_t.grid(row=3, column=1, padx=5, pady=5)
      self.delta_1_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.delta_1_t.grid(row=4, column=1, padx=5, pady=5)
      self.caudal_1_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.caudal_1_t.grid(row=5, column=1, padx=5, pady=5)

      Label(frame0, text='Sensor 2', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=2, padx=5, pady=5, sticky='nsew')
      self.valor_actual_2_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.valor_actual_2_t.grid(row=1, column=2, padx=5, pady=5)
      self.tiempo_2_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_2_t.grid(row=2, column=2, padx=5, pady=5)
      self.tiempo_anterior_2_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_anterior_2_t.grid(row=3, column=2, padx=5, pady=5)
      self.delta_2_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.delta_2_t.grid(row=4, column=2, padx=5, pady=5)
      self.caudal_2_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.caudal_2_t.grid(row=5, column=2, padx=5, pady=5)

      Label(frame0, text='Sensor 3', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=3, padx=5, pady=5, sticky='nsew')
      self.valor_actual_3_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.valor_actual_3_t.grid(row=1, column=3, padx=5, pady=5)
      self.tiempo_3_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_3_t.grid(row=2, column=3, padx=5, pady=5)
      self.tiempo_anterior_3_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_anterior_3_t.grid(row=3, column=3, padx=5, pady=5)
      self.delta_3_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.delta_3_t.grid(row=4, column=3, padx=5, pady=5)
      self.caudal_3_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.caudal_3_t.grid(row=5, column=3, padx=5, pady=5)

      Label(frame0, text='Sensor 4', bg='#090808', fg='white', font=('Arial', 12, 'bold')).grid(row=0, column=4, padx=5, pady=5, sticky='nsew')
      self.valor_actual_4_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.valor_actual_4_t.grid(row=1, column=4, padx=5, pady=5)
      self.tiempo_4_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_4_t.grid(row=2, column=4, padx=5, pady=5)
      self.tiempo_anterior_4_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.tiempo_anterior_4_t.grid(row=3, column=4, padx=5, pady=5)
      self.delta_4_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.delta_4_t.grid(row=4, column=4, padx=5, pady=5)
      self.caudal_4_t = Label(frame0, text="none", font=('Arial', 12, 'bold'))
      self.caudal_4_t.grid(row=5, column=4, padx=5, pady=5)

      ## configuracion relativa del los tamaños dentro del frame 0
      frame0.columnconfigure(0,weight=1)
      frame0.columnconfigure(1,weight=2)
      frame0.columnconfigure(2,weight=2)
      frame0.columnconfigure(3,weight=2)
      frame0.columnconfigure(4,weight=2)
      frame0.rowconfigure(0, weight=1)
      frame0.rowconfigure(1, weight=1)
      frame0.rowconfigure(2, weight=1)
      frame0.rowconfigure(3, weight=1)
      frame0.rowconfigure(4, weight=1)
      frame0.rowconfigure(5, weight=1)
   
      ## define  botones
      self.bt_iniciar = Button(frame, text='Iniciar', font=('Arial', 12, 'bold'),
                              width=12, bg='#0cbccc', fg='black', command=self.funcion_a_realizar)
      self.bt_iniciar.pack(pady=5, expand=1)

      self.bt_guardar = Button(frame1, text='Guardar', font=('Arial', 12, 'bold'),
                              width=12, bg='#2898ee', fg='black', command=self.guardar_datos)
      self.bt_guardar.pack(pady=5, expand=1)

      self.bt_conectar = Button(frame3, text='Conectar', font=('Arial', 12, 'bold'),
                              width=12, bg='#284eee', fg='black', command=self.conectar_serial)
      self.bt_conectar.pack(pady=5, expand=1)

      self.bt_desconectar = Button(frame5, state='disabled', text='Desconectar', font=('Arial', 12, 'bold'),
                              width=12, bg='#284eee', fg='black', command=self.desconectar_serial)
      self.bt_desconectar.pack(pady=5, expand=1)

      ## extrae informacion de puertos y velocidades
      baud = self.datos_placa.baudrates

      self.combobox_baud = ttk.Combobox(frame6, values=baud, justify='center', width=12, font='Arial')
      self.combobox_baud.pack(padx=20, expand=1)
      self.combobox_baud.current(3)
      try:
         port = self.datos_placa.puertos

         self.combobox_port = ttk.Combobox(frame4, values=port, justify='center', width=12, font='Arial')
         self.combobox_port.pack(pady=0, expand=1)
         self.combobox_port.current(0)
      except:
         pass

      ## Crea un hilo para majar la parte logica del almacenamiento de los datos
      self.CrearHilo()

      
   def funcion_a_realizar(self):

      print("hola")
      
      segundos = float(self.delta_aux.total_seconds())
      print(type(segundos))
      print(segundos)


   def guardar_datos(self):

      self.sheet.title = self.nombre_documento.get() # Cambiar el nombre de la hoja de cálculo
      self.libro.save(filename= self.sheet.title+'.xlsx')# Guardar el libro de trabajo en un archivo

   def conectar_serial(self):

      self.datos_placa.placa.port = self.combobox_port.get()
      self.datos_placa.placa.baudrate = self.combobox_baud.get()
      self.datos_placa.conexion_serial()
      if (self.datos_placa.placa.is_open):

         # cambia de estado los botones
         self.bt_conectar.config(state='disabled')
         self.bt_desconectar.config(state='normal')

         # Toma los valores de las constantes y los bloquea
         self.factor[0] = float(self.factor_1.get())
         self.constante_1_e.config(state='disabled')
         self.factor[1] = float(self.factor_2.get())
         self.constante_2_e.config(state='disabled')

   def desconectar_serial(self):
      self.bt_conectar.config(state='normal')
      self.bt_desconectar.config(state='disabled')
      try:
         self.ani.event_source.stop()
      except AttributeError:
         pass
      self.datos_placa.desconectar()

      # Desbloquea las constantes
      self.constante_1_e.config(state='normal')
      self.constante_2_e.config(state='normal')

## Inicia el bucle principal
if __name__ == "__main__":
   #archivo_texto = open('datos_uart.txt', 'w')
   terminal = Tk()
   terminal.geometry('742x535')
   terminal.config(bg='#010808', bd=4)
   terminal.wm_title('Caudalimetro')
   terminal.minsize(width=700, height=400)  # Corregido el nombre de 'minisize' a 'minsize'
   #ventana.call('wm', 'iconphoto', ventana._w, PhotoImage(file='logo.png'))
   app = Grafica(terminal)
   app.mainloop()
